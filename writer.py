import pandas as pd
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.utils.dataframe import dataframe_to_rows

FONT = {'size': 14, 'bold': True}

class Writer:

    def __init__(self, book, cursor, writer):
        self.book = book
        self.cursor = cursor
        self.writer = writer

    def execute_queries(self, first_row, queries, columns, sheet_name, table_names):
        current_row = first_row
        query_index = 0
        for query in queries:
            last_row = self.write(current_row, query, query_index, columns, sheet_name, table_names)
            current_row = last_row + 6
            query_index += 1

    def execute_queries_with_plot(self, first_row, queries, columns, sheet_name, table_names):
        current_row = first_row
        query_index = 0
        for query in queries:
            last_row = self.write_with_plot(current_row, query, query_index, columns, sheet_name, table_names)
            current_row = last_row + 6
            query_index += 1

    def write(self, first_row, query, query_index, columns, sheet_name, table_names):
        self.cursor.execute(query)
        data = self.cursor.fetchall()
        last_row = first_row + len(data) + 1
        # Create Pandas dataframe from the data
        df = pd.DataFrame(data, columns=columns)
        # Convert dataframe to XlsxWriter Excel object
        # df.to_excel(self.writer, sheet_name=sheet_name, startrow=current_row)
        worksheet = self.book[sheet_name]
        # worksheet = self.writer.sheets[sheet_name]
        # df[columns[1]] = pd.to_numeric(df[columns[1]], errors='coerce')
        df[columns[1]] = df[columns[1]].astype(float)
        current_row = first_row
        # with pd.option_context('display.max_rows', None, 'display.max_columns', None):  # more options can be specified also
        #     print(df)
        for r in dataframe_to_rows(df, index=False, header=True):
            worksheet.cell(row=current_row, column=1).value = r[0]
            worksheet.cell(row=current_row, column=2).value = r[1]
            # print(worksheet[current_row])
            # worksheet[current_row] = r
            # worksheet.append(r)
            current_row += 1
        # Increase column width for date
        # if columns[0] == 'Datum':
            # worksheet.set_column(1, 1, 10)
        # Add table title
        # worksheet.write_string(current_row - 1, 0, table_names[query_index])
        worksheet.cell(row=first_row - 1, column=1, value=table_names[query_index])
        return last_row

    def write_with_plot(self, first_row, query, query_index, columns, sheet_name, table_names):
        self.cursor.execute(query)
        data = self.cursor.fetchall()
        # first_row = current_row + 2
        last_row = first_row + len(data)
        # Create Pandas dataframe from the data
        df = pd.DataFrame(data, columns=columns)
        # Convert dataframe to XlsxWriter Excel object
        # df.to_excel(self.writer, sheet_name=sheet_name, startrow=current_row)
        # worksheet = self.writer.sheets[sheet_name]
        worksheet = self.book[sheet_name]
        df[columns[1]] = df[columns[1]].astype(float)
        current_row = first_row
        # df[columns[1]] = pd.to_numeric(df[columns[1]], errors='coerce')
        for r in dataframe_to_rows(df, index=False, header=True):
            worksheet.cell(row=current_row, column=1).value = r[0]
            worksheet.cell(row=current_row, column=2).value = r[1]
            # print(worksheet[current_row])
            # worksheet[current_row] = r
            # worksheet.append(r)
            current_row += 1
        # Increase column width for date
        # if columns[0] == 'Datum':
            # worksheet.set_column(1, 1, 10)
        # Add table title
        # worksheet.write_string(current_row - 1, 0, table_names[query_index])
        worksheet.cell(row=first_row - 1, column=1, value=table_names[query_index])
        # Add colouring to values
        # worksheet.conditional_format('C{}:C{}'.format(first_row, last_row), {'type': '3_color_scale'})
        self.create_chart(worksheet, data, query_index, columns[0], first_row, last_row, sheet_name, table_names)
        return last_row


    def create_chart(self, worksheet, data, query_index, column, first_row, last_row, sheet_name, table_names):
        # Create a chart object
        # workbook  = self.writer.book
        # chart = workbook.add_chart({'type': 'line'})
        chart = LineChart()
        chart.title = "Line Chart"
        chart.style = 13
        chart.y_axis.title = 'Size'
        chart.x_axis.title = 'Test Number'
        # chart.set_size({'width': 900, 'height': 550})
        # chart.set_y_axis({
        #     'name': column,
        #     'name_font': FONT
        # })
        # chart.set_x_axis({
        #     'name': 'Datum',
        #     'name_font': FONT
        # })
        # chart.set_title({ 'name': table_names[query_index]})
        # chart.set_legend({'none': True})
        # chart.set_style(37)
        # Configure the series of the chart from the data
        # y_values = '={}!$C${}:$C${}'.format(sheet_name, first_row, last_row)
        # x_values = '={}!$B${}:$B${}'.format(sheet_name, first_row, last_row)
        # chart.add_series({'name': column, 'values': y_values, 'categories': x_values})

        # Insert the chart into the worksheet.
        # worksheet.insert_chart('F{}'.format(first_row), chart)
        data = Reference(worksheet, min_col=2, min_row=first_row + 1, max_col=2, max_row=last_row)
        chart.add_data(data)
        # series = Series(data, title="First series of values")
        # chart.append(series)
        worksheet.add_chart(chart, 'F{}'.format(first_row))